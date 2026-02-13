from __future__ import annotations

import csv
import json
import subprocess
import sys
from collections import Counter
from pathlib import Path

import pytest
from openpyxl import Workbook

from equitable_lunch_invites.cli import main as cli_main
from equitable_lunch_invites.io import load_guest_roster, load_host_roster, normalize_outcome
from equitable_lunch_invites.models import DEMOGRAPHIC_MODE_PROPORTIONAL, DEMOGRAPHIC_MODE_WOMEN_TO_PARITY
from equitable_lunch_invites.models import DISCIPLINES
from equitable_lunch_invites.selection import apply_demographic_mode, demographic_targets_for_event


def run_cli(args: list[str]) -> None:
    exit_code = cli_main(args)
    assert exit_code == 0


def make_workbook(path: Path) -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = "host_roster"
    ws.append(["Name", "Discipline"])
    ws.append(["Host A", "biophysics"])
    ws.append(["Host B", "astronomy"])
    ws.append(["Host C", "condensed matter experimental"])
    ws.append(["Host D", "condensed matter theory"])
    ws.append(["Host E", "biophysics"])

    ws = wb.create_sheet("guest_roster")
    ws.append(["Name", "Discipline", "Sex"])
    ws.append(["Guest A", "biophysics", "F"])
    ws.append(["Guest B", "astronomy", "M"])
    ws.append(["Guest C", "condensed matter experimental", "F"])
    ws.append(["Guest D", "condensed matter theory", "M"])
    ws.append(["Guest E", "biophysics", "F"])
    ws.append(["Guest F", "astronomy", "M"])
    ws.append(["Guest G", "condensed matter experimental", "F"])
    ws.append(["Guest H", "condensed matter theory", "M"])
    wb.save(path)


def read_plan_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        return [{key: (value or "") for key, value in row.items()} for row in reader]


def write_plan_rows(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def test_init_defaults_to_data_paths_and_seed_key(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    workbook = tmp_path / "planner_inputs.xlsx"
    make_workbook(workbook)
    monkeypatch.chdir(tmp_path)

    run_cli(["init", "--inputs", str(workbook), "--seed-key", "42"])

    state_path = tmp_path / "data" / "state.json"
    plan_csv = tmp_path / "data" / "lunch_plan.csv"
    assert state_path.exists()
    assert plan_csv.exists()

    state = json.loads(state_path.read_text(encoding="utf-8"))
    assert state["_meta"]["seed_key"] == 42
    assert isinstance(state["_meta"]["generated_seed"], int)


def test_plan_writes_single_csv_for_selected_and_waitlist(tmp_path: Path) -> None:
    workbook = tmp_path / "planner_inputs.xlsx"
    make_workbook(workbook)
    state_path = tmp_path / "data" / "state.json"
    plan_csv = tmp_path / "data" / "lunch_plan.csv"

    run_cli(
        [
            "init",
            "--inputs",
            str(workbook),
            "--seed-key",
            "7",
            "--state",
            str(state_path),
            "--plan-csv",
            str(plan_csv),
        ]
    )
    run_cli(
        [
            "plan",
            "--inputs",
            str(workbook),
            "--state",
            str(state_path),
            "--plan-csv",
            str(plan_csv),
            "--event-index",
            "1",
            "--hosts-per-event",
            "2",
            "--guests-per-event",
            "3",
        ]
    )

    rows = read_plan_rows(plan_csv)
    assert rows
    assert {row["event_index"] for row in rows} == {"1"}
    assert {"host", "guest"} <= {row["role"] for row in rows}
    assert {"selected", "waitlist"} <= {row["status"] for row in rows}
    assert all("attendance" in row for row in rows)

    selected_guests = [
        row for row in rows if row["role"] == "guest" and row["status"] == "selected"
    ]
    assert len(selected_guests) == 3


def test_next_plan_applies_previous_attendance_from_same_csv(tmp_path: Path) -> None:
    workbook = tmp_path / "planner_inputs.xlsx"
    make_workbook(workbook)
    state_path = tmp_path / "data" / "state.json"
    plan_csv = tmp_path / "data" / "lunch_plan.csv"

    run_cli(
        [
            "init",
            "--inputs",
            str(workbook),
            "--seed-key",
            "99",
            "--state",
            str(state_path),
            "--plan-csv",
            str(plan_csv),
        ]
    )
    run_cli(
        [
            "plan",
            "--inputs",
            str(workbook),
            "--state",
            str(state_path),
            "--plan-csv",
            str(plan_csv),
            "--event-index",
            "1",
            "--hosts-per-event",
            "1",
            "--guests-per-event",
            "3",
        ]
    )

    rows = read_plan_rows(plan_csv)
    event_1_selected_guests = [
        row
        for row in rows
        if row["event_index"] == "1"
        and row["role"] == "guest"
        and row["status"] == "selected"
    ]
    assert len(event_1_selected_guests) == 3

    for idx, row in enumerate(event_1_selected_guests):
        row["attendance"] = "no_show" if idx == 0 else "attended"
    by_key = {
        (row["event_index"], row["role"], row["status"], row["name"]): row
        for row in rows
    }
    for row in event_1_selected_guests:
        by_key[(row["event_index"], row["role"], row["status"], row["name"])] = row
    write_plan_rows(plan_csv, list(by_key.values()))

    run_cli(
        [
            "plan",
            "--inputs",
            str(workbook),
            "--state",
            str(state_path),
            "--plan-csv",
            str(plan_csv),
            "--hosts-per-event",
            "1",
            "--guests-per-event",
            "3",
        ]
    )

    state = json.loads(state_path.read_text(encoding="utf-8"))
    no_show_guest = event_1_selected_guests[0]["name"]
    assert state["guests"][no_show_guest]["assigned_count"] == 1
    assert state["guests"][no_show_guest]["no_show_count"] == 1
    assert state["_meta"]["last_recorded_event"] == 1
    assert state["_meta"]["last_planned_event"] == 2

    rows_after = read_plan_rows(plan_csv)
    assert {"1", "2"} <= {row["event_index"] for row in rows_after}


def test_cant_attend_and_waitlist_filled_are_applied(tmp_path: Path) -> None:
    workbook = tmp_path / "planner_inputs.xlsx"
    make_workbook(workbook)
    state_path = tmp_path / "data" / "state.json"
    plan_csv = tmp_path / "data" / "lunch_plan.csv"

    run_cli(
        [
            "init",
            "--inputs",
            str(workbook),
            "--seed-key",
            "61",
            "--state",
            str(state_path),
            "--plan-csv",
            str(plan_csv),
        ]
    )
    run_cli(
        [
            "plan",
            "--inputs",
            str(workbook),
            "--state",
            str(state_path),
            "--plan-csv",
            str(plan_csv),
            "--event-index",
            "1",
            "--hosts-per-event",
            "1",
            "--guests-per-event",
            "3",
        ]
    )

    rows = read_plan_rows(plan_csv)
    selected_rows = [
        row
        for row in rows
        if row["event_index"] == "1"
        and row["role"] == "guest"
        and row["status"] == "selected"
    ]
    waitlist_rows = [
        row
        for row in rows
        if row["event_index"] == "1"
        and row["role"] == "guest"
        and row["status"] == "waitlist"
    ]
    assert len(selected_rows) == 3
    assert waitlist_rows

    selected_rows[0]["attendance"] = "cant_attend"
    selected_rows[1]["attendance"] = "attended"
    selected_rows[2]["attendance"] = "no_show"
    waitlist_rows[0]["attendance"] = "filled"
    if len(waitlist_rows) > 1:
        waitlist_rows[1]["attendance"] = "cant_attend"

    by_key = {
        (row["event_index"], row["role"], row["status"], row["name"]): row
        for row in rows
    }
    for row in selected_rows + waitlist_rows:
        key = (row["event_index"], row["role"], row["status"], row["name"])
        by_key[key] = row
    write_plan_rows(plan_csv, list(by_key.values()))

    run_cli(
        [
            "plan",
            "--inputs",
            str(workbook),
            "--state",
            str(state_path),
            "--plan-csv",
            str(plan_csv),
            "--hosts-per-event",
            "1",
            "--guests-per-event",
            "3",
        ]
    )

    state = json.loads(state_path.read_text(encoding="utf-8"))
    cant_attend_name = selected_rows[0]["name"]
    attended_name = selected_rows[1]["name"]
    no_show_name = selected_rows[2]["name"]
    replacement_name = waitlist_rows[0]["name"]

    assert state["guests"][cant_attend_name]["assigned_count"] == 0
    assert state["guests"][cant_attend_name]["no_show_count"] == 0
    assert state["guests"][cant_attend_name]["cooldown"] == 0
    assert state["guests"][attended_name]["assigned_count"] == 1
    assert state["guests"][no_show_name]["assigned_count"] == 1
    assert state["guests"][no_show_name]["no_show_count"] == 1
    assert state["guests"][no_show_name]["cooldown"] == 1
    assert state["guests"][replacement_name]["assigned_count"] == 1
    assert state["guests"][replacement_name]["no_show_count"] == 0
    assert state["guests"][replacement_name]["cooldown"] == 0


def test_normalize_outcome_accepts_extended_attendance_values() -> None:
    assert normalize_outcome("can't attend") == "cant_attend"
    assert normalize_outcome("declined") == "cant_attend"
    assert normalize_outcome("filled in") == "filled"
    assert normalize_outcome("replacement") == "filled"


def test_plan_requires_attendance_for_prior_event(tmp_path: Path) -> None:
    workbook = tmp_path / "planner_inputs.xlsx"
    make_workbook(workbook)
    state_path = tmp_path / "data" / "state.json"
    plan_csv = tmp_path / "data" / "lunch_plan.csv"

    run_cli(
        [
            "init",
            "--inputs",
            str(workbook),
            "--seed-key",
            "101",
            "--state",
            str(state_path),
            "--plan-csv",
            str(plan_csv),
        ]
    )
    run_cli(
        [
            "plan",
            "--inputs",
            str(workbook),
            "--state",
            str(state_path),
            "--plan-csv",
            str(plan_csv),
            "--event-index",
            "1",
            "--hosts-per-event",
            "1",
            "--guests-per-event",
            "2",
        ]
    )

    with pytest.raises(ValueError, match="Missing guest attendance"):
        cli_main(
            [
                "plan",
                "--inputs",
                str(workbook),
                "--state",
                str(state_path),
                "--plan-csv",
                str(plan_csv),
                "--event-index",
                "2",
                "--hosts-per-event",
                "1",
                "--guests-per-event",
                "2",
            ]
        )


def test_plan_excludes_hosts_from_guest_selection(tmp_path: Path) -> None:
    workbook = tmp_path / "planner_inputs.xlsx"
    wb = Workbook()

    ws = wb.active
    ws.title = "host_roster"
    ws.append(["Name", "Discipline"])
    ws.append(["Shared Name", "biophysics"])
    ws.append(["Host B", "astronomy"])

    ws = wb.create_sheet("guest_roster")
    ws.append(["Name", "Discipline", "Sex"])
    ws.append(["Shared Name", "biophysics", "F"])
    ws.append(["Guest A", "astronomy", "M"])
    ws.append(["Guest B", "condensed matter experimental", "F"])
    ws.append(["Guest C", "condensed matter theory", "M"])
    wb.save(workbook)

    state_path = tmp_path / "data" / "state.json"
    plan_csv = tmp_path / "data" / "lunch_plan.csv"

    run_cli(
        [
            "start",
            "--inputs",
            str(workbook),
            "--seed-key",
            "88",
            "--state",
            str(state_path),
            "--plan-csv",
            str(plan_csv),
            "--event-index",
            "1",
            "--hosts-per-event",
            "1",
            "--guests-per-event",
            "2",
        ]
    )

    rows = read_plan_rows(plan_csv)
    guest_names = {
        row["name"]
        for row in rows
        if row["event_index"] == "1" and row["role"] == "guest"
    }
    assert "Shared Name" not in guest_names


def test_seed_key_must_be_bounded() -> None:
    with pytest.raises(SystemExit):
        cli_main(["init", "--inputs", "x.xlsx", "--seed-key", "1001"])


def test_missing_disciplines_are_split_equally(tmp_path: Path) -> None:
    host_path = tmp_path / "hosts.csv"
    with host_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["Name", "Discipline"])
        for idx in range(8):
            writer.writerow([f"Host {idx}", ""])
    hosts = load_host_roster(host_path)
    host_counts = Counter(participant.discipline for participant in hosts)
    assert set(host_counts.keys()) == set(DISCIPLINES)
    assert sorted(host_counts.values()) == [2, 2, 2, 2]

    guest_path = tmp_path / "guests.csv"
    with guest_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["Name", "Discipline", "Sex"])
        writer.writerow(["Guest Explicit", "astronomy", "F"])
        for idx in range(9):
            writer.writerow([f"Guest {idx}", "", "M"])
    guests = load_guest_roster(guest_path)
    guest_counts = Counter(participant.discipline for participant in guests)
    assert guest_counts["astronomy"] >= 1
    assert max(guest_counts.values()) - min(guest_counts.values()) <= 1


def test_start_initializes_and_plans_in_one_command(tmp_path: Path) -> None:
    workbook = tmp_path / "planner_inputs.xlsx"
    make_workbook(workbook)
    state_path = tmp_path / "data" / "state.json"
    plan_csv = tmp_path / "data" / "lunch_plan.csv"

    run_cli(
        [
            "start",
            "--inputs",
            str(workbook),
            "--seed-key",
            "77",
            "--state",
            str(state_path),
            "--plan-csv",
            str(plan_csv),
            "--event-index",
            "1",
            "--hosts-per-event",
            "1",
            "--guests-per-event",
            "2",
        ]
    )

    assert state_path.exists()
    assert plan_csv.exists()
    state = json.loads(state_path.read_text(encoding="utf-8"))
    assert state["_meta"]["seed_key"] == 77
    assert state["_meta"]["last_planned_event"] == 1
    rows = read_plan_rows(plan_csv)
    assert any(
        row["event_index"] == "1"
        and row["role"] == "guest"
        and row["status"] == "selected"
        for row in rows
    )


def test_women_to_parity_mode_targets_reach_parity_over_time() -> None:
    counts = {"F": 8, "M": 52}

    parity_weights = apply_demographic_mode(
        demographic_counts=counts,
        demographic_column="Sex",
        demographic_mode=DEMOGRAPHIC_MODE_WOMEN_TO_PARITY,
    )
    proportional_weights = apply_demographic_mode(
        demographic_counts=counts,
        demographic_column="Sex",
        demographic_mode=DEMOGRAPHIC_MODE_PROPORTIONAL,
    )

    assert parity_weights["F"] == parity_weights["M"] == 52
    assert proportional_weights == counts

    parity_cumulative: Counter[str] = Counter()
    proportional_cumulative: Counter[str] = Counter()
    for event_index in (1, 2, 3, 4):
        parity_cumulative.update(
            demographic_targets_for_event(event_index, 5, parity_weights)
        )
        proportional_cumulative.update(
            demographic_targets_for_event(event_index, 5, proportional_weights)
        )

    assert parity_cumulative["F"] == parity_cumulative["M"]
    assert proportional_cumulative["F"] < proportional_cumulative["M"]


def test_legacy_cli_still_works(tmp_path: Path) -> None:
    roster = tmp_path / "legacy_roster.csv"
    state_path = tmp_path / "legacy_state.json"
    with roster.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["Name", "Sex", "Discipline"])
        writer.writerow(["Legacy A", "F", "biophysics"])
        writer.writerow(["Legacy B", "M", "astronomy"])
        writer.writerow(["Legacy C", "F", "condensed matter experimental"])
        writer.writerow(["Legacy D", "M", "condensed matter theory"])

    repo_root = Path(__file__).resolve().parents[1]
    subprocess.run(
        [sys.executable, "lunches.py", "init", "--roster", str(roster), "--state", str(state_path)],
        cwd=repo_root,
        check=True,
    )
    selected = tmp_path / "legacy_selected.csv"
    waitlist = tmp_path / "legacy_wait.csv"
    subprocess.run(
        [
            sys.executable,
            "lunches.py",
            "plan",
            "--roster",
            str(roster),
            "--state",
            str(state_path),
            "--lunch-index",
            "1",
            "--seed",
            "100",
            "--out",
            str(selected),
            "--wait",
            str(waitlist),
            "--seats",
            "2",
            "--max-unique",
            "3",
        ],
        cwd=repo_root,
        check=True,
    )
    assert selected.exists()
    assert waitlist.exists()
