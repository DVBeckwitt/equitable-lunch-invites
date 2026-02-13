from __future__ import annotations

import csv
import math
from pathlib import Path
from typing import Iterable

from equitable_lunch_invites.models import DISCIPLINE_ALIASES, DISCIPLINES, Participant

ANONYMOUS_GUEST_DISCIPLINE = "guest pool"
PLAN_CSV_HEADER = [
    "event_index",
    "role",
    "status",
    "name",
    "discipline",
    "demographic",
    "attendance",
]
PLAN_ROLE_HOST = "host"
PLAN_ROLE_GUEST = "guest"
PLAN_STATUS_SELECTED = "selected"
PLAN_STATUS_WAITLIST = "waitlist"


def round_half_up(value: float) -> int:
    return int(math.floor(value + 0.5))


def normalize_header(value: str) -> str:
    return " ".join((value or "").strip().lower().split())


def normalize_discipline(value: str) -> str:
    token = normalize_header(value)
    return DISCIPLINE_ALIASES.get(token, token)


def normalize_sex(value: str) -> str:
    token_upper = (value or "").strip().upper()
    if token_upper in {"F", "M"}:
        return token_upper
    token = normalize_header(value)
    if token in {"female", "woman", "w"}:
        return "F"
    if token in {"male", "man"}:
        return "M"
    return "U"


def normalize_demographic(value: str) -> str:
    token = " ".join((value or "").strip().split())
    if not token:
        return "U"
    if len(token) <= 3:
        return token.upper()
    return token


def normalize_outcome(value: str) -> str:
    token = normalize_header((value or "").replace("_", " ").replace("-", " "))
    if not token:
        return ""

    attended_tokens = {
        "attended",
        "attend",
        "present",
        "showed",
        "showed up",
        "yes",
        "y",
        "a",
    }
    no_show_tokens = {
        "no show",
        "noshow",
        "missed",
        "absent",
        "no",
        "n",
        "ns",
    }
    cant_attend_tokens = {
        "cant attend",
        "can't attend",
        "cannot attend",
        "declined",
        "decline",
        "not available",
        "unavailable",
    }
    filled_tokens = {
        "filled",
        "filled in",
        "fill in",
        "replacement",
        "waitlist filled",
    }

    if token in attended_tokens:
        return "attended"
    if token in no_show_tokens:
        return "no_show"
    if token in cant_attend_tokens:
        return "cant_attend"
    if token in filled_tokens:
        return "filled"
    raise ValueError(
        f"Unknown guest outcome value '{value}'. "
        "Use attended/no_show/cant_attend/filled (or leave blank)."
    )


def _read_csv_dict(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        fieldnames = [name for name in (reader.fieldnames or []) if name]
        rows = [{k: (v or "") for k, v in row.items()} for row in reader]
    return fieldnames, rows


def _read_xlsx_dict(path: Path, sheet_name: str | None = None) -> tuple[list[str], list[dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Reading .xlsx files requires openpyxl. Install with: pip install openpyxl"
        ) from exc

    workbook = load_workbook(path, data_only=True, read_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found in {path}. Available sheets: {workbook.sheetnames}"
            )
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook[workbook.sheetnames[0]]

    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return [], []

    header_idx = None
    for idx, row in enumerate(rows):
        if row and any((str(cell).strip() if cell is not None else "") for cell in row):
            header_idx = idx
            break

    if header_idx is None:
        return [], []

    header_raw = rows[header_idx]
    fieldnames = [
        str(cell).strip()
        for cell in header_raw
        if cell is not None and str(cell).strip()
    ]
    if not fieldnames:
        return [], []

    parsed_rows: list[dict[str, str]] = []
    width = len(fieldnames)
    for row in rows[header_idx + 1 :]:
        values = list(row) if row is not None else []
        if len(values) < width:
            values.extend([None] * (width - len(values)))
        trimmed = values[:width]
        if not any((str(cell).strip() if cell is not None else "") for cell in trimmed):
            continue
        parsed_rows.append(
            {
                fieldnames[i]: ("" if trimmed[i] is None else str(trimmed[i]).strip())
                for i in range(width)
            }
        )
    return fieldnames, parsed_rows


def read_table_dict(path: Path, sheet_name: str | None = None) -> tuple[list[str], list[dict[str, str]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv_dict(path)
    if suffix == ".xlsx":
        return _read_xlsx_dict(path, sheet_name=sheet_name)
    raise ValueError(f"Unsupported file type '{path.suffix}'. Use .csv or .xlsx")


def _column_map(fieldnames: list[str]) -> dict[str, str]:
    return {normalize_header(name): name for name in fieldnames}


def _validate_columns(column_map: dict[str, str], required: list[str], label: str) -> None:
    missing = [name for name in required if name not in column_map]
    if missing:
        raise ValueError(
            f"{label} CSV is missing required columns: {missing}. "
            f"Expected case-insensitive headers."
        )


def _normalize_role(value: str) -> str:
    token = normalize_header(value)
    aliases = {
        "host": "host",
        "hosts": "host",
        "h": "host",
        "guest": "guest",
        "guests": "guest",
        "g": "guest",
    }
    return aliases.get(token, token)


def _ensure_unique_names(participants: Iterable[Participant], label: str) -> None:
    seen: set[str] = set()
    duplicates: set[str] = set()
    for participant in participants:
        if participant.name in seen:
            duplicates.add(participant.name)
        seen.add(participant.name)
    if duplicates:
        preview = sorted(duplicates)[:10]
        raise ValueError(f"Duplicate names in {label} roster: {preview}")


def _distribute_missing_disciplines_evenly(raw_disciplines: list[str]) -> list[str]:
    distributed = list(raw_disciplines)
    missing_indices = [index for index, value in enumerate(raw_disciplines) if not value.strip()]
    if not missing_indices:
        return distributed

    for offset, row_index in enumerate(missing_indices):
        distributed[row_index] = DISCIPLINES[offset % len(DISCIPLINES)]
    return distributed


def load_host_roster(path: Path, sheet_name: str | None = None) -> list[Participant]:
    fieldnames, rows = read_table_dict(path, sheet_name=sheet_name)
    if not fieldnames:
        raise ValueError("Host roster file is empty or missing headers.")
    columns = _column_map(fieldnames)
    _validate_columns(columns, ["name", "discipline"], "Host roster")

    prepared: list[tuple[str, str]] = []
    for row in rows:
        name = (row.get(columns["name"]) or "").strip()
        discipline = normalize_discipline(row.get(columns["discipline"]) or "")
        if not name:
            continue
        if discipline and discipline not in DISCIPLINES:
            raise ValueError(
                f"Unknown discipline '{discipline}' for host '{name}'. "
                f"Allowed disciplines: {DISCIPLINES}"
            )
        prepared.append((name, discipline))

    distributed_disciplines = _distribute_missing_disciplines_evenly(
        [discipline for _, discipline in prepared]
    )
    out = [
        Participant(name=name, discipline=distributed_disciplines[index])
        for index, (name, _) in enumerate(prepared)
    ]

    _ensure_unique_names(out, "host")
    if not out:
        raise ValueError("Host roster has no valid rows.")
    return out


def load_guest_roster(
    path: Path,
    demographic_column: str = "Sex",
    outcome_column: str | None = None,
    sheet_name: str | None = None,
) -> list[Participant]:
    fieldnames, rows = read_table_dict(path, sheet_name=sheet_name)
    if not fieldnames:
        raise ValueError("Guest roster file is empty or missing headers.")
    columns = _column_map(fieldnames)

    demographic_key = normalize_header(demographic_column)
    required_columns = ["name", "discipline", demographic_key]
    outcome_key = normalize_header(outcome_column) if outcome_column else None
    if outcome_key:
        required_columns.append(outcome_key)
    _validate_columns(columns, required_columns, "Guest roster")

    prepared: list[tuple[str, str, str]] = []
    for row in rows:
        name = (row.get(columns["name"]) or "").strip()
        discipline = normalize_discipline(row.get(columns["discipline"]) or "")
        raw_demographic = row.get(columns[demographic_key]) or ""
        demographic = normalize_sex(raw_demographic) if demographic_key == "sex" else normalize_demographic(raw_demographic)
        if not name:
            continue
        if discipline and discipline not in DISCIPLINES:
            raise ValueError(
                f"Unknown discipline '{discipline}' for guest '{name}'. "
                f"Allowed disciplines: {DISCIPLINES}"
            )
        if outcome_key:
            normalize_outcome(row.get(columns[outcome_key]) or "")
        prepared.append((name, discipline, demographic))

    distributed_disciplines = _distribute_missing_disciplines_evenly(
        [discipline for _, discipline, _ in prepared]
    )
    out = [
        Participant(
            name=name,
            discipline=distributed_disciplines[index],
            demographic=demographic,
        )
        for index, (name, _, demographic) in enumerate(prepared)
    ]

    _ensure_unique_names(out, "guest")
    if not out:
        raise ValueError("Guest roster has no valid rows.")
    return out


def _normalize_guest_discipline(value: str) -> str:
    if not value.strip():
        return ANONYMOUS_GUEST_DISCIPLINE
    discipline = normalize_discipline(value)
    if discipline == ANONYMOUS_GUEST_DISCIPLINE:
        return discipline
    if discipline not in DISCIPLINES:
        raise ValueError(
            f"Unknown guest discipline '{discipline}'. "
            f"Allowed disciplines: {DISCIPLINES + [ANONYMOUS_GUEST_DISCIPLINE]}"
        )
    return discipline


def _parse_optional_count(raw_value: str, row_index: int) -> int:
    token = (raw_value or "").strip()
    if not token:
        return 0
    try:
        value = int(token)
    except ValueError as exc:
        raise ValueError(f"Invalid Count at row {row_index}: '{raw_value}'") from exc
    if value < 0:
        raise ValueError(f"Count cannot be negative at row {row_index}.")
    return value


def load_merged_roster(
    path: Path,
    demographic_column: str = "Sex",
    sheet_name: str | None = None,
) -> tuple[list[Participant], list[Participant], bool]:
    fieldnames, rows = read_table_dict(path, sheet_name=sheet_name)
    if not fieldnames:
        raise ValueError("Merged roster file is empty or missing headers.")
    columns = _column_map(fieldnames)
    _validate_columns(columns, ["role"], "Merged roster")

    demographic_key = normalize_header(demographic_column)
    has_demographic = demographic_key in columns
    count_column = columns.get("count")
    name_column = columns.get("name")
    discipline_column = columns.get("discipline")

    if discipline_column is None:
        raise ValueError("Merged roster CSV must include a Discipline column.")

    hosts: list[Participant] = []
    named_guests: list[Participant] = []
    anonymous_guest_specs: list[tuple[str, str, int]] = []

    for idx, row in enumerate(rows, start=2):
        role_raw = row.get(columns["role"]) or ""
        role = _normalize_role(role_raw)
        if not role:
            continue
        if role not in {"host", "guest"}:
            raise ValueError(
                f"Unknown role '{role_raw}' at row {idx}. Use host or guest."
            )

        name = (row.get(name_column) or "").strip() if name_column else ""
        discipline_raw = row.get(discipline_column) or ""
        count = _parse_optional_count(row.get(count_column) or "", idx) if count_column else 0
        raw_demographic = row.get(columns[demographic_key]) if has_demographic else ""
        raw_demographic = raw_demographic or ""
        demographic = normalize_sex(raw_demographic) if demographic_key == "sex" else normalize_demographic(raw_demographic)

        if role == "host":
            if not name:
                raise ValueError(f"Host rows require Name (row {idx}).")
            if count > 0:
                raise ValueError(f"Host rows cannot use Count (row {idx}).")
            discipline = normalize_discipline(discipline_raw)
            if discipline not in DISCIPLINES:
                raise ValueError(
                    f"Unknown discipline '{discipline}' for host '{name}'. "
                    f"Allowed disciplines: {DISCIPLINES}"
                )
            hosts.append(Participant(name=name, discipline=discipline))
            continue

        # Guest role
        if name:
            if count > 0:
                raise ValueError(
                    f"Guest row {idx} cannot set both Name and Count."
                )
            discipline = _normalize_guest_discipline(discipline_raw)
            named_guests.append(
                Participant(name=name, discipline=discipline, demographic=demographic)
            )
        else:
            if count <= 0:
                raise ValueError(
                    f"Guest row {idx} requires Name or positive Count."
                )
            discipline = _normalize_guest_discipline(discipline_raw)
            anonymous_guest_specs.append((discipline, demographic, count))

    if not hosts:
        raise ValueError("Merged roster has no valid host rows.")

    generated_guests: list[Participant] = []
    used_names = {participant.name for participant in named_guests}
    next_number = 1
    for discipline, demographic, count in anonymous_guest_specs:
        for _ in range(count):
            while str(next_number) in used_names:
                next_number += 1
            generated_name = str(next_number)
            generated_guests.append(
                Participant(
                    name=generated_name,
                    discipline=discipline,
                    demographic=demographic,
                )
            )
            used_names.add(generated_name)
            next_number += 1

    guests = named_guests + generated_guests
    _ensure_unique_names(hosts, "host")
    _ensure_unique_names(guests, "guest")
    if not guests:
        raise ValueError("Merged roster has no valid guest rows.")

    guest_is_anonymous = bool(generated_guests) and not bool(named_guests)
    return hosts, guests, guest_is_anonymous


def build_numbered_guests(total_guests: int) -> list[Participant]:
    if total_guests <= 0:
        raise ValueError("--guest-total must be >= 1.")
    return [
        Participant(
            name=str(number),
            discipline=ANONYMOUS_GUEST_DISCIPLINE,
            demographic="U",
        )
        for number in range(1, total_guests + 1)
    ]


def read_guest_outcomes(
    path: Path,
    outcome_column: str = "Outcome",
    sheet_name: str | None = None,
) -> dict[str, str]:
    fieldnames, rows = read_table_dict(path, sheet_name=sheet_name)
    if not fieldnames:
        raise ValueError("Guest roster file is empty or missing headers.")
    columns = _column_map(fieldnames)
    outcome_key = normalize_header(outcome_column)
    _validate_columns(columns, ["name", outcome_key], "Guest roster")

    statuses: dict[str, str] = {}
    seen_names: set[str] = set()
    for row in rows:
        name = (row.get(columns["name"]) or "").strip()
        if not name:
            continue
        if name in seen_names:
            raise ValueError(f"Duplicate guest name in roster outcomes: {name}")
        seen_names.add(name)

        raw_status = row.get(columns[outcome_key]) or ""
        normalized = normalize_outcome(raw_status)
        if normalized:
            statuses[name] = normalized
    return statuses


def read_firstcol(path: Path, sheet_name: str | None = None) -> list[str]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", newline="", encoding="utf-8-sig") as handle:
            reader = csv.reader(handle)
            rows = [row for row in reader if row and any(cell.strip() for cell in row)]
        if not rows:
            return []
        first_cell = rows[0][0].strip().lower()
        start = 1 if first_cell in {"name", "student", "student name"} else 0
        return [row[0].strip() for row in rows[start:] if row and row[0].strip()]

    fieldnames, rows = read_table_dict(path, sheet_name=sheet_name)
    if not fieldnames:
        return []
    first_column = fieldnames[0]
    header_token = first_column.strip().lower()
    names = [row.get(first_column, "").strip() for row in rows if row.get(first_column, "").strip()]
    if header_token in {"name", "student", "student name"}:
        return names
    return names


def read_csv_firstcol(path: Path) -> list[str]:
    # Backward-compatible alias for legacy callers.
    return read_firstcol(path)


def write_participant_list(
    path: Path,
    participants: list[Participant],
    demographic_header: str | None = None,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    header = ["name", "discipline"]
    normalized_demo_header = None
    if demographic_header:
        normalized_demo_header = demographic_header.strip().lower()
        header.append(normalized_demo_header)

    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(header)
        for participant in participants:
            row = [participant.name, participant.discipline]
            if normalized_demo_header:
                row.append(participant.demographic or "")
            writer.writerow(row)


def _normalize_plan_role(value: str) -> str:
    token = normalize_header(value)
    if token in {"host", "hosts"}:
        return PLAN_ROLE_HOST
    if token in {"guest", "guests"}:
        return PLAN_ROLE_GUEST
    raise ValueError(f"Unknown plan CSV role '{value}'. Use host or guest.")


def _normalize_plan_status(value: str) -> str:
    token = normalize_header(value)
    if token == PLAN_STATUS_SELECTED:
        return PLAN_STATUS_SELECTED
    if token == PLAN_STATUS_WAITLIST:
        return PLAN_STATUS_WAITLIST
    raise ValueError(f"Unknown plan CSV status '{value}'. Use selected or waitlist.")


def _parse_event_index(raw_value: str) -> int:
    token = (raw_value or "").strip()
    try:
        value = int(token)
    except ValueError as exc:
        raise ValueError(f"Invalid event_index '{raw_value}' in plan CSV.") from exc
    if value <= 0:
        raise ValueError(f"event_index must be >= 1 in plan CSV, got '{raw_value}'.")
    return value


def _read_existing_plan_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        fieldnames = [name for name in (reader.fieldnames or []) if name]
        if not fieldnames:
            return []

        columns = _column_map(fieldnames)
        _validate_columns(
            columns,
            [normalize_header(name) for name in PLAN_CSV_HEADER],
            "Plan",
        )
        rows = [{key: (value or "").strip() for key, value in row.items()} for row in reader]
    return rows


def read_plan_attendance(
    path: Path,
    event_index: int,
    role: str,
    status: str = PLAN_STATUS_SELECTED,
) -> dict[str, str]:
    normalized_role = _normalize_plan_role(role)
    normalized_status = _normalize_plan_status(status)
    target_event = int(event_index)
    if target_event <= 0:
        raise ValueError("event_index must be >= 1.")

    rows = _read_existing_plan_rows(path)
    attendance_by_name: dict[str, str] = {}
    for row in rows:
        row_event = _parse_event_index(row.get("event_index", ""))
        if row_event != target_event:
            continue
        if _normalize_plan_role(row.get("role", "")) != normalized_role:
            continue
        if _normalize_plan_status(row.get("status", "")) != normalized_status:
            continue

        name = (row.get("name") or "").strip()
        if not name:
            continue
        if name in attendance_by_name:
            raise ValueError(
                f"Duplicate {normalized_role} '{name}' in event {target_event} plan rows."
            )
        attendance_by_name[name] = normalize_outcome(row.get("attendance") or "")
    return attendance_by_name


def _participant_rows(
    event_index: int,
    role: str,
    status: str,
    participants: list[Participant],
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for participant in participants:
        rows.append(
            {
                "event_index": str(event_index),
                "role": role,
                "status": status,
                "name": participant.name,
                "discipline": participant.discipline,
                "demographic": participant.demographic or "",
                "attendance": "",
            }
        )
    return rows


def upsert_event_plan_csv(
    path: Path,
    event_index: int,
    selected_hosts: list[Participant],
    waitlist_hosts: list[Participant],
    selected_guests: list[Participant],
    waitlist_guests: list[Participant],
) -> None:
    target_event = int(event_index)
    if target_event <= 0:
        raise ValueError("event_index must be >= 1.")

    existing_rows = _read_existing_plan_rows(path)
    previous_attendance: dict[tuple[str, str, str], str] = {}
    kept_rows: list[dict[str, str]] = []
    for row in existing_rows:
        row_event = _parse_event_index(row.get("event_index", ""))
        role = _normalize_plan_role(row.get("role", ""))
        status = _normalize_plan_status(row.get("status", ""))
        name = (row.get("name") or "").strip()
        if row_event == target_event and name:
            previous_attendance[(role, status, name)] = normalize_outcome(row.get("attendance") or "")
            continue
        kept_rows.append(
            {
                "event_index": str(row_event),
                "role": role,
                "status": status,
                "name": name,
                "discipline": (row.get("discipline") or "").strip(),
                "demographic": (row.get("demographic") or "").strip(),
                "attendance": normalize_outcome(row.get("attendance") or ""),
            }
        )

    new_rows = (
        _participant_rows(target_event, PLAN_ROLE_HOST, PLAN_STATUS_SELECTED, selected_hosts)
        + _participant_rows(target_event, PLAN_ROLE_HOST, PLAN_STATUS_WAITLIST, waitlist_hosts)
        + _participant_rows(target_event, PLAN_ROLE_GUEST, PLAN_STATUS_SELECTED, selected_guests)
        + _participant_rows(target_event, PLAN_ROLE_GUEST, PLAN_STATUS_WAITLIST, waitlist_guests)
    )
    for row in new_rows:
        key = (row["role"], row["status"], row["name"])
        row["attendance"] = previous_attendance.get(key, "")

    all_rows = kept_rows + new_rows
    status_order = {PLAN_STATUS_SELECTED: 0, PLAN_STATUS_WAITLIST: 1}
    role_order = {PLAN_ROLE_HOST: 0, PLAN_ROLE_GUEST: 1}
    all_rows.sort(
        key=lambda row: (
            _parse_event_index(row.get("event_index", "")),
            status_order.get(row.get("status", ""), 99),
            role_order.get(row.get("role", ""), 99),
            (row.get("name") or "").casefold(),
        )
    )

    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=PLAN_CSV_HEADER)
        writer.writeheader()
        writer.writerows(all_rows)
